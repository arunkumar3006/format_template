"""
api/data_loader.py (Lightweight Version)
---------------------------------------
A pandas-free implementation using pure openpyxl and csv.
This avoids the 400MB bundle size issue on Vercel.
"""

import csv
import io
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Canonical columns and their variations
COLUMN_ALIASES = {
    "title": ["title", "headline", "subject", "article_title", "news_title"],
    "author_or_journalist": ["author", "journalist", "reporter", "writer", "byline", "author_name", "author/reporter", "author/journalist"],
    "publisher_name": ["publisher", "publication", "source", "publisher_name", "media_house", "publisher/agency", "publisher/agency"],
    "date_time": ["date", "time", "date_time", "published_at", "timestamp", "published at"],
    "summary_of_article": ["summary", "description", "content", "abstract", "summary_of_article"],
    "link": ["link", "url", "url_of_article", "source_url", "article_url", "resolved url", "resolved_url", "link_of_article"],
    "category": ["category", "section", "news_type", "tag", "industry"]
}

def _normalise(val) -> str:
    """Safely return a clean string."""
    if val is None: return "N/A"
    s = str(val).strip()
    return s if s and s.lower() not in ("nan", "none", "n/a") else "N/A"

def _match_column(header_name: str) -> str:
    """Match a raw header against the canonical aliases."""
    if not header_name: return "unknown"
    # Aggressive normalization: remove non-alphanumeric for matching
    clean = "".join(c for c in str(header_name).lower() if c.isalnum() or c == "_")
    norm_header = clean.strip().replace(" ", "_")
    
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            clean_alias = "".join(c for c in alias.lower() if c.isalnum() or c == "_")
            if norm_header == clean_alias:
                return canonical
    return header_name  # Keep as-is if no match

def load_dataset(file_object, source_filename: str = "") -> tuple[list, str]:
    """
    Load data as a list of dictionaries (Mocking a DataFrame-like list).
    Supports .xlsx, .xls (partially), and .csv.
    """
    data = []
    try:
        raw = file_object.read()
        
        filename = source_filename.lower() if source_filename else getattr(file_object, "name", "upload").lower()
        
        logger.info(f"DATA LOADER: Parsing file: source_filename='{source_filename}', evaluated filename='{filename}'")

        # ─── CASE: CSV ────────────────────────────────────────────────────────
        if filename.endswith(".csv") or filename == "upload":
            stream = io.StringIO(raw.decode('utf-8', errors='ignore'))
            reader = csv.DictReader(stream)
            headers = {h: _match_column(h) for h in (reader.fieldnames or [])}
            for row in reader:
                data.append({headers[k]: _normalise(v) for k, v in row.items()})

        # ─── CASE: XLSX ───────────────────────────────────────────────────────
        else:
            # Use read_only=True for speed and memory efficiency
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            
            rows = ws.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                return [], "❌ The sheet is empty."

            # Map headers
            col_map = {idx: _match_column(str(val)) for idx, val in enumerate(header_row) if val is not None}
            
            for row in rows:
                if not any(row): continue
                record = {}
                for idx, val in enumerate(row):
                    if idx in col_map:
                        record[col_map[idx]] = _normalise(val)
                data.append(record)

    except Exception as exc:
        logger.exception("Failed to load dataset.")
        return [], f"❌ Load error: {exc}"

    if not data:
        return [], "❌ No data found."

    # Final cleanup: ensure all canonical columns exist in each dict
    for row in data:
        for col in COLUMN_ALIASES.keys():
            if col not in row:
                row[col] = "N/A"

    msg = f"✅ Loaded {len(data):,} articles (Lightweight Mode)."
    logger.info(msg)
    return data, msg
